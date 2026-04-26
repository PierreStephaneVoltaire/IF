"""Excel export builder for training programs.

Uses openpyxl to produce a multi-sheet .xlsx workbook from a program dict.
"""
from __future__ import annotations

import copy
import json
import os
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Shared styles
_HEADER_FONT = Font(bold=True, size=11)
_SECTION_FONT = Font(bold=True, size=11, color="1F3864")
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="E7ECF5", end_color="E7ECF5", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")

# openpyxl cell content limit
_CELL_CHAR_LIMIT = 32000


def _num(v: Any) -> Any:
    """Coerce a value to an Excel-compatible scalar."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (dict, list)):
        return str(v)
    return v


def _is_blank(v: Any) -> bool:
    return v is None or v == ""


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value[:10]).date()
    except ValueError:
        return None


def _extract_phase_name(phase: Any) -> str:
    """Extract phase name from either a string or a phase dict."""
    if isinstance(phase, dict):
        return phase.get("name", "")
    return phase if isinstance(phase, str) else str(phase)


def _truncate(val: Any, limit: int = _CELL_CHAR_LIMIT) -> Any:
    if isinstance(val, str) and len(val) > limit:
        return val[: limit - 30] + "... [truncated]"
    return val


def _serialize_json(value: Any) -> str:
    if value is None:
        return ""
    try:
        return json.dumps(value, default=str, ensure_ascii=True, sort_keys=True)
    except TypeError:
        return str(value)


def _first_non_blank(*values: Any) -> Any:
    for value in values:
        if not _is_blank(value):
            return value
    return ""


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], col_widths: dict[int, int] | None = None):
    """Write a header row + data rows to a worksheet."""
    # Header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=_truncate(_num(value)))

    # Column widths
    max_col = len(headers)
    for col_idx in range(1, max_col + 1):
        if col_widths and col_idx in col_widths:
            width = col_widths[col_idx]
        else:
            # Auto-width from header + data
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            width = min(max_len + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"


def _write_section_label(ws, row_idx: int, label: str, width: int = 1):
    """Write a bold, filled section label row starting at column A."""
    cell = ws.cell(row=row_idx, column=1, value=label)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    if width > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=width)


def _autosize_columns(ws, max_cols: int, min_width: int = 10, max_width: int = 60):
    """Size columns based on longest value in any row."""
    for col_idx in range(1, max_cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            s = str(val)
            longest_line = max((len(line) for line in s.splitlines()), default=len(s))
            if longest_line > max_len:
                max_len = longest_line
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _fmt(val: Any) -> Any:
    """Format a value for a key-value cell: join lists, coerce numbers, truncate."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, list):
        return "; ".join(str(_num(v)) for v in val)
    return _truncate(_num(val))


def _flatten_rows(value: Any, prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key in sorted(value.keys(), key=str):
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_rows(value[key], next_prefix))
        return rows
    if isinstance(value, list):
        for idx, item in enumerate(value):
            next_prefix = f"{prefix}[{idx}]"
            rows.extend(_flatten_rows(item, next_prefix))
        return rows
    rows.append((prefix or "(root)", _truncate(_fmt(value))))
    return rows


def _write_kv_sheet(ws, rows: list[tuple[str, Any]], section_width: int = 2) -> None:
    _write_sheet(ws, ["Field", "Value"], [[k, v] for k, v in rows], col_widths={1: 34, 2: 72})
    ws.freeze_panes = "A2"


def _write_raw_sheet(wb: Workbook, title: str, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(title)
    headers = ["Source", "Path", "Value"]
    rows: list[list[Any]] = []
    for source, value in payload.items():
        for path, cell_value in _flatten_rows(value, prefix=""):
            rows.append([source, path, cell_value])
    if not rows:
        rows.append(["—", "(empty)", ""])
    _write_sheet(ws, headers, rows, col_widths={1: 18, 2: 52, 3: 72})


def _safe_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if value is None:
        return []
    return [value]


def _deepcopy_rows(rows: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    return [copy.deepcopy(row) for row in (rows or [])]


def _bodyweight_candidates(sessions: list[dict[str, Any]], idx: int) -> tuple[tuple[Any, str] | None, tuple[Any, str] | None]:
    prev: tuple[Any, str] | None = None
    next_: tuple[Any, str] | None = None

    for offset in range(1, len(sessions)):
        left = idx - offset
        right = idx + offset
        if prev is None and left >= 0:
            candidate = sessions[left].get("body_weight_kg")
            if not _is_blank(candidate):
                prev = (candidate, "previous session")
        if next_ is None and right < len(sessions):
            candidate = sessions[right].get("body_weight_kg")
            if not _is_blank(candidate):
                next_ = (candidate, "next session")
        if prev is not None and next_ is not None:
            break

    return prev, next_


def _nearest_weight_log_value(
    target_date: date | None,
    weight_log: list[dict[str, Any]],
) -> Any:
    if target_date is None or not weight_log:
        return ""

    best: tuple[int, Any] | None = None
    for entry in weight_log:
        entry_date = _parse_date(str(entry.get("date", "")))
        if entry_date is None or _is_blank(entry.get("kg")):
            continue
        delta = abs((entry_date - target_date).days)
        if best is None or delta < best[0]:
            best = (delta, entry.get("kg"))
    return best[1] if best is not None else ""


def _resolve_bodyweight_for_session(
    sessions: list[dict[str, Any]],
    idx: int,
    weight_log: list[dict[str, Any]],
    meta: dict[str, Any],
) -> tuple[Any, str]:
    session = sessions[idx]
    direct = session.get("body_weight_kg")
    if not _is_blank(direct):
        return _num(direct), "session"

    prev, next_ = _bodyweight_candidates(sessions, idx)
    if prev and next_:
        prev_distance = 0
        next_distance = 0
        # Index distance is sufficient because sessions are exported in date order.
        for offset in range(1, len(sessions)):
            if idx - offset >= 0 and not _is_blank(sessions[idx - offset].get("body_weight_kg")):
                prev_distance = offset
                break
        for offset in range(1, len(sessions)):
            if idx + offset < len(sessions) and not _is_blank(sessions[idx + offset].get("body_weight_kg")):
                next_distance = offset
                break
        chosen = prev if prev_distance <= next_distance else next_
        return _num(chosen[0]), chosen[1]
    if prev:
        return _num(prev[0]), prev[1]
    if next_:
        return _num(next_[0]), next_[1]

    session_date = _parse_date(str(session.get("date", "")))
    from_log = _nearest_weight_log_value(session_date, weight_log)
    if not _is_blank(from_log):
        return _num(from_log), "weight log"

    meta_bw = _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))
    if not _is_blank(meta_bw):
        return _num(meta_bw), "meta"

    return 0.0, "fallback"


def _resolve_bodyweight_for_date(
    target_date: str | None,
    sessions: list[dict[str, Any]],
    weight_log: list[dict[str, Any]],
    meta: dict[str, Any],
) -> tuple[Any, str]:
    parsed_target = _parse_date(target_date)
    if parsed_target is None:
        meta_bw = _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))
        return (_num(meta_bw) if not _is_blank(meta_bw) else 0.0, "meta" if not _is_blank(meta_bw) else "fallback")

    direct_candidates: list[tuple[int, Any, str]] = []
    for session in sessions:
        if _is_blank(session.get("body_weight_kg")):
            continue
        session_date = _parse_date(str(session.get("date", "")))
        if session_date is None:
            continue
        direct_candidates.append((abs((session_date - parsed_target).days), session.get("body_weight_kg"), "session"))
    if direct_candidates:
        direct_candidates.sort(key=lambda item: item[0])
        return _num(direct_candidates[0][1]), direct_candidates[0][2]

    log_value = _nearest_weight_log_value(parsed_target, weight_log)
    if not _is_blank(log_value):
        return _num(log_value), "weight log"

    meta_bw = _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))
    if not _is_blank(meta_bw):
        return _num(meta_bw), "meta"

    return 0.0, "fallback"


def build_program_xlsx(
    program: dict,
    out_path: str,
    analysis: dict | None = None,
    export_context: dict[str, Any] | None = None,
) -> str:
    """Build a multi-sheet Excel workbook from a program dict.

    Args:
        program: Full program dict from ProgramStore.
        out_path: Absolute path to write the .xlsx file.
        analysis: Optional analysis bundle with keys:
            - "weekly": output of analytics.weekly_analysis()
            - "correlation": cached correlation report dict (or None if not generated)
            - "program_evaluation": cached program evaluation dict (or None)
            - "lift_profiles": normalized lift profiles list (from summarize_lift_profiles)
            When provided, appends analysis sheets.
        export_context: Optional sidecar payload with keys such as:
            - pk: active program partition key / creator fallback
            - version: active program version token, e.g. "v003"
            - weight_log: list of weight log entries
            - max_history: list of max history entries
            - glossary: exercise glossary rows
            - federation_library: shared federation and standards record
            - sex: preferred sex for DOTS calculations

    Returns:
        Absolute path to the written file.
    """
    wb = Workbook()
    export_context = export_context or {}

    meta = copy.deepcopy(program.get("meta", {}))
    phases = _deepcopy_rows(program.get("phases"))
    sessions = _deepcopy_rows(program.get("sessions"))
    goals = _deepcopy_rows(program.get("goals"))
    competitions = _deepcopy_rows(program.get("competitions"))
    diet_notes = _deepcopy_rows(program.get("diet_notes"))
    supplements = _deepcopy_rows(program.get("supplements"))
    supplement_phases = _deepcopy_rows(program.get("supplement_phases"))
    lift_profiles = _deepcopy_rows(program.get("lift_profiles"))
    breaks = _deepcopy_rows(program.get("breaks"))
    current_maxes = copy.deepcopy(program.get("current_maxes", {}))
    operator_prefs = copy.deepcopy(program.get("operator_prefs", {}))

    weight_log = _deepcopy_rows(export_context.get("weight_log"))
    max_history = _deepcopy_rows(export_context.get("max_history"))
    glossary = _deepcopy_rows(export_context.get("glossary"))
    federation_library = copy.deepcopy(export_context.get("federation_library") or {})
    creator = _first_non_blank(meta.get("creator"), export_context.get("pk"), program.get("pk"), "operator")
    program_pk = _first_non_blank(export_context.get("pk"), program.get("pk"), creator)
    sex = str(_first_non_blank(meta.get("sex"), export_context.get("sex"), "male")).lower()
    version_token = str(_first_non_blank(export_context.get("version"), meta.get("version"), meta.get("version_label"), "")).strip()
    if version_token and not version_token.startswith("v") and version_token.isdigit():
        version_token = f"v{int(version_token):03d}"

    meta.setdefault("sex", sex)
    meta.setdefault("creator", creator)
    if version_token:
        meta.setdefault("program_version", version_token)

    prepared_sessions = _prepare_sessions_for_export(sessions, weight_log, meta)

    # ---- Base program sheets ----
    _write_meta_sheet(wb, meta, creator=creator, program_pk=program_pk, version_token=version_token, sex=sex)
    _write_goals_sheet(wb, meta, goals, competitions, sex, federation_library)
    _write_federation_standards_sheet(wb, federation_library)
    _write_current_maxes_sheet(wb, current_maxes)
    _write_max_history_sheet(wb, max_history, sex, meta)
    _write_phases_sheet(wb, phases)
    _write_sessions_sheet(wb, prepared_sessions)
    _write_planned_exercises_sheet(wb, prepared_sessions)
    _write_exercises_sheet(wb, prepared_sessions)
    _write_competitions_sheet(wb, competitions, prepared_sessions, weight_log, sex, meta, federation_library)
    _write_biometrics_sheet(wb, diet_notes)
    _write_weight_log_sheet(wb, weight_log)
    _write_supplements_sheet(wb, supplements)
    _write_supplement_phases_sheet(wb, supplement_phases)
    _write_lift_profiles_sheet(wb, lift_profiles)
    _write_exercise_glossary_sheet(wb, glossary)
    _write_videos_sheet(wb, prepared_sessions)
    _write_notes_sheet(wb, meta, phases, prepared_sessions, competitions, diet_notes, supplement_phases)
    _write_breaks_and_prefs_sheet(wb, breaks, operator_prefs, meta)

    # ---- Analysis sheets ----
    if analysis is not None:
        _write_weekly_analysis_sheet(wb, analysis.get("weekly") or {})
        _write_trends_sheet(wb, prepared_sessions, weight_log, diet_notes, sex, meta)
        _write_per_lift_metrics_sheet(wb, analysis.get("weekly") or {})
        _write_correlation_sheet(wb, analysis.get("correlation"))
        _write_program_evaluation_sheet(wb, analysis.get("program_evaluation"))
    else:
        _write_weekly_analysis_sheet(wb, {})
        _write_trends_sheet(wb, prepared_sessions, weight_log, diet_notes, sex, meta)
        _write_per_lift_metrics_sheet(wb, {})
        _write_correlation_sheet(wb, None)
        _write_program_evaluation_sheet(wb, None)

    # ---- Raw backup ----
    _write_raw_sheet(
        wb,
        "Raw Export",
        {
            "program": program,
            "analysis": analysis or {},
            "weight_log": weight_log,
            "max_history": max_history,
            "glossary": glossary,
            "federation_library": federation_library,
        },
    )

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Analysis sheet builders
# ---------------------------------------------------------------------------

_MISSING_NOTE = "Not yet generated — open the Analysis page and let the report render before exporting."


def _write_lift_profiles_sheet(wb: Workbook, lift_profiles: list[dict]) -> None:
    ws = wb.create_sheet("Lift Profiles")
    headers = [
        "Lift",
        "Style Notes",
        "Sticking Points",
        "Primary Muscle",
        "Volume Tolerance",
        "Stimulus Coefficient",
        "Coefficient Reasoning",
        "Confidence",
        "Updated At",
    ]
    rows = []
    for p in lift_profiles:
        rows.append([
            (p.get("lift") or "").title(),
            p.get("style_notes") or "",
            p.get("sticking_points") or "",
            p.get("primary_muscle") or "",
            p.get("volume_tolerance") or "",
            p.get("stimulus_coefficient", ""),
            p.get("stimulus_coefficient_reasoning", ""),
            p.get("stimulus_coefficient_confidence", ""),
            p.get("stimulus_coefficient_updated_at", ""),
        ])
    if not rows:
        rows.append(["—", "No lift profiles recorded on this program.", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={2: 40, 3: 40, 7: 36})


def _write_weekly_analysis_sheet(wb: Workbook, weekly: dict) -> None:
    ws = wb.create_sheet("Weekly Analysis")
    if not weekly:
        _write_sheet(ws, ["Field", "Value"], [["Status", _MISSING_NOTE]])
        return

    compliance = weekly.get("compliance") or {}
    deload = weekly.get("deload_info") or {}
    current_maxes = weekly.get("current_maxes") or {}
    readiness = weekly.get("readiness_score") or {}
    flags = weekly.get("flags") or []
    bodyweight_trend = weekly.get("bodyweight_trend") or {}
    exercise_stats = weekly.get("exercise_stats") or {}
    fatigue_components = weekly.get("fatigue_components") or {}
    fatigue_dimensions = weekly.get("fatigue_dimensions") or {}
    banister = weekly.get("banister") or {}
    monotony_strain = weekly.get("monotony_strain") or {}
    decoupling = weekly.get("decoupling") or {}
    taper_quality = weekly.get("taper_quality") or {}
    specificity = weekly.get("specificity_ratio") or {}
    inol = weekly.get("inol") or {}
    acwr = weekly.get("acwr") or {}
    volume_landmarks = weekly.get("volume_landmarks") or {}
    peaking_timeline = weekly.get("peaking_timeline") or {}
    projection_calibration = weekly.get("projection_calibration") or {}
    alerts = weekly.get("alerts") or []

    rows = [
        ["Current Week", weekly.get("week", "")],
        ["Block", weekly.get("block", "")],
        ["Sessions Analyzed", weekly.get("sessions_analyzed", "")],
        ["Compliance — Planned", compliance.get("planned", "")],
        ["Compliance — Completed", compliance.get("completed", "")],
        ["Compliance — %", compliance.get("pct", "")],
        ["Fatigue Index", weekly.get("fatigue_index", "")],
        ["Estimated DOTS", weekly.get("estimated_dots", "")],
        ["Current Max — Squat (kg)", current_maxes.get("squat", "")],
        ["Current Max — Bench (kg)", current_maxes.get("bench", "")],
        ["Current Max — Deadlift (kg)", current_maxes.get("deadlift", "")],
        ["Current Maxes Method", current_maxes.get("method", "")],
        ["Deload Weeks", _fmt(deload.get("deload_weeks"))],
        ["Break Weeks", _fmt(deload.get("break_weeks"))],
        ["Effective Training Weeks", deload.get("effective_training_weeks", "")],
        ["Readiness Score", readiness.get("score", "") if isinstance(readiness, dict) else ""],
        ["Readiness Zone", readiness.get("zone", "") if isinstance(readiness, dict) else ""],
        ["Readiness - Fatigue", readiness.get("components", {}).get("fatigue_norm", "") if isinstance(readiness, dict) else ""],
        ["Readiness - RPE Drift", readiness.get("components", {}).get("rpe_drift", "") if isinstance(readiness, dict) else ""],
        ["Readiness - Wellness", readiness.get("components", {}).get("wellness", "") if isinstance(readiness, dict) else ""],
        ["Readiness - Trend", readiness.get("components", {}).get("performance_trend", "") if isinstance(readiness, dict) else ""],
        ["Readiness - BW Deviation", readiness.get("components", {}).get("bw_deviation", "") if isinstance(readiness, dict) else ""],
        ["Exercise Stats", _serialize_json(exercise_stats)],
        ["Fatigue Components", _serialize_json(fatigue_components)],
        ["Fatigue Dimensions", _serialize_json(fatigue_dimensions)],
        ["Banister", _serialize_json(banister)],
        ["Monotony / Strain", _serialize_json(monotony_strain)],
        ["Decoupling", _serialize_json(decoupling)],
        ["Taper Quality", _serialize_json(taper_quality)],
        ["Specificity Ratio", _serialize_json(specificity)],
        ["INOL", _serialize_json(inol)],
        ["ACWR", _serialize_json(acwr)],
        ["Volume Landmarks", _serialize_json(volume_landmarks)],
        ["Peaking Timeline", _serialize_json(peaking_timeline)],
        ["Projection Calibration", _serialize_json(projection_calibration)],
        ["Bodyweight Trend", _serialize_json(bodyweight_trend)],
        ["Alerts", _serialize_json(alerts)],
        ["Flags", _fmt(flags)],
    ]
    _write_sheet(ws, ["Field", "Value"], rows, col_widths={1: 32, 2: 60})


def _write_per_lift_metrics_sheet(wb: Workbook, weekly: dict) -> None:
    ws = wb.create_sheet("Per-Lift Metrics")

    lifts = (weekly or {}).get("lifts") or {}
    projections = (weekly or {}).get("projections") or []

    headers = [
        "Lift", "Progression (kg/wk)", "Fit Quality", "Kendall τ", "Volume Change %",
        "Intensity Change %", "Failed Sets", "RPE Trend",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP

    row_idx = 2
    if lifts:
        for name, data in lifts.items():
            values = [
                name,
                _fmt(data.get("progression_rate_kg_per_week")),
                _fmt(data.get("fit_quality", data.get("r_squared", data.get("r2")))),
                _fmt(data.get("kendall_tau")),
                _fmt(data.get("volume_change_pct")),
                _fmt(data.get("intensity_change_pct")),
                _fmt(data.get("failed_sets")),
                _fmt(data.get("rpe_trend")),
            ]
            for col_idx, v in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
    else:
        ws.cell(row=row_idx, column=1, value="No tracked lifts in the window.")
        row_idx += 1

    # Projections sub-section
    row_idx += 1
    _write_section_label(ws, row_idx, "Projections", width=len(headers))
    row_idx += 1
    proj_headers = ["Comp", "Weeks Out", "Projected Total (kg)", "Confidence", "Method"]
    for col_idx, h in enumerate(proj_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    row_idx += 1

    if projections:
        for p in projections:
            values = [
                p.get("comp_name") or "(unscheduled)",
                _fmt(p.get("weeks_to_comp")),
                _fmt(p.get("total")),
                _fmt(p.get("confidence")),
                _fmt(p.get("method")),
            ]
            for col_idx, v in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
    else:
        reason = (weekly or {}).get("projection_reason") or "No upcoming competitions to project."
        ws.cell(row=row_idx, column=1, value=reason)

    _autosize_columns(ws, max_cols=len(headers))
    ws.freeze_panes = "A2"


def _write_correlation_sheet(wb: Workbook, correlation: dict | None) -> None:
    ws = wb.create_sheet("ROI Correlation")

    if not correlation:
        _write_sheet(ws, ["Field", "Value"], [["Status", _MISSING_NOTE]])
        return

    # Header block
    header_rows = [
        ["Summary", _fmt(correlation.get("summary"))],
        ["Generated At", _fmt(correlation.get("generated_at"))],
        ["Window Start", _fmt(correlation.get("window_start"))],
        ["Weeks", _fmt(correlation.get("weeks"))],
        ["Cached", _fmt(correlation.get("cached"))],
    ]
    for row_idx, (k, v) in enumerate(header_rows, 1):
        c1 = ws.cell(row=row_idx, column=1, value=k)
        c1.font = _HEADER_FONT
        c1.fill = _HEADER_FILL
        c1.alignment = _WRAP
        c2 = ws.cell(row=row_idx, column=2, value=v)
        c2.alignment = _WRAP

    next_row = len(header_rows) + 2

    if correlation.get("insufficient_data"):
        ws.cell(row=next_row, column=1, value="Insufficient data").font = _SECTION_FONT
        ws.cell(row=next_row, column=2, value=_fmt(correlation.get("insufficient_data_reason")))
        _autosize_columns(ws, max_cols=6)
        return

    _write_section_label(ws, next_row, "Findings", width=6)
    next_row += 1

    finding_headers = ["Exercise", "Lift", "Direction", "Strength", "Reasoning", "Caveat"]
    for col_idx, h in enumerate(finding_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    next_row += 1

    findings = correlation.get("findings") or []
    if findings:
        for f in findings:
            values = [
                _fmt(f.get("exercise")),
                _fmt(f.get("lift")),
                _fmt(f.get("correlation_direction")),
                _fmt(f.get("strength")),
                _fmt(f.get("reasoning")),
                _fmt(f.get("caveat")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="No findings reported.")

    _autosize_columns(ws, max_cols=6)


def _write_program_evaluation_sheet(wb: Workbook, evaluation: dict | None) -> None:
    ws = wb.create_sheet("Program Evaluation")

    if not evaluation:
        _write_sheet(ws, ["Field", "Value"], [["Status", _MISSING_NOTE]])
        return

    header_rows = [
        ["Stance", _fmt(evaluation.get("stance"))],
        ["Summary", _fmt(evaluation.get("summary"))],
        ["Conclusion", _fmt(evaluation.get("conclusion"))],
        ["Generated At", _fmt(evaluation.get("generated_at"))],
        ["Window Start", _fmt(evaluation.get("window_start"))],
        ["Weeks", _fmt(evaluation.get("weeks"))],
        ["Cached", _fmt(evaluation.get("cached"))],
    ]
    for row_idx, (k, v) in enumerate(header_rows, 1):
        c1 = ws.cell(row=row_idx, column=1, value=k)
        c1.font = _HEADER_FONT
        c1.fill = _HEADER_FILL
        c1.alignment = _WRAP
        c2 = ws.cell(row=row_idx, column=2, value=v)
        c2.alignment = _WRAP

    next_row = len(header_rows) + 2

    if evaluation.get("insufficient_data"):
        ws.cell(row=next_row, column=1, value="Insufficient data").font = _SECTION_FONT
        ws.cell(row=next_row, column=2, value=_fmt(evaluation.get("insufficient_data_reason")))
        _autosize_columns(ws, max_cols=4)
        return

    # String-list sections
    for label, key in (("What's Working", "what_is_working"),
                       ("What's Not Working", "what_is_not_working"),
                       ("Monitoring Focus", "monitoring_focus")):
        _write_section_label(ws, next_row, label, width=4)
        next_row += 1
        items = evaluation.get(key) or []
        if items:
            for item in items:
                c = ws.cell(row=next_row, column=1, value=_truncate(str(item)))
                c.alignment = _WRAP
                ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                next_row += 1
        else:
            ws.cell(row=next_row, column=1, value="(none)")
            next_row += 1
        next_row += 1

    # Competition Alignment
    _write_section_label(ws, next_row, "Competition Alignment", width=4)
    next_row += 1
    ca_headers = ["Competition", "Role", "Alignment", "Reason"]
    for col_idx, h in enumerate(ca_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    next_row += 1
    alignment_items = evaluation.get("competition_alignment") or []
    if alignment_items:
        for item in alignment_items:
            values = [
                _fmt(item.get("competition")),
                _fmt(item.get("role")),
                _fmt(item.get("alignment")),
                _fmt(item.get("reason")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="(none)")
        next_row += 1
    next_row += 1

    _write_section_label(ws, next_row, "Goal Status", width=4)
    next_row += 1
    goal_headers = ["Goal", "Priority", "Status", "Reason"]
    for col_idx, h in enumerate(goal_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    next_row += 1
    goal_items = evaluation.get("goal_status") or []
    if goal_items:
        for item in goal_items:
            values = [
                _fmt(item.get("goal")),
                _fmt(item.get("priority")),
                _fmt(item.get("status")),
                _fmt(item.get("reason")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="(none)")
        next_row += 1
    next_row += 1

    _write_section_label(ws, next_row, "Competition Strategy", width=4)
    next_row += 1
    strategy_headers = ["Competition", "Priority", "Approach", "Reason"]
    for col_idx, h in enumerate(strategy_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    next_row += 1
    strategy_items = evaluation.get("competition_strategy") or []
    if strategy_items:
        for item in strategy_items:
            values = [
                _fmt(item.get("competition")),
                _fmt(item.get("priority")),
                _fmt(item.get("approach")),
                _fmt(item.get("reason")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="(none)")
        next_row += 1
    next_row += 1

    _write_section_label(ws, next_row, "Weight Class Strategy", width=4)
    next_row += 1
    weight_strategy = evaluation.get("weight_class_strategy") or {}
    weight_rows = [
        ["Recommendation", _fmt(weight_strategy.get("recommendation"))],
        ["Recommended Weight Class", _fmt(weight_strategy.get("recommended_weight_class_kg"))],
    ]
    for label, value in weight_rows:
        ws.cell(row=next_row, column=1, value=label).font = _HEADER_FONT
        ws.cell(row=next_row, column=1).fill = _HEADER_FILL
        ws.cell(row=next_row, column=1).alignment = _WRAP
        ws.cell(row=next_row, column=2, value=value).alignment = _WRAP
        next_row += 1
    option_headers = ["Option Weight Class", "Suitability", "Reason"]
    for col_idx, h in enumerate(option_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    next_row += 1
    viable_options = weight_strategy.get("viable_options") or []
    if viable_options:
        for item in viable_options:
            values = [
                _fmt(item.get("weight_class_kg")),
                _fmt(item.get("suitability")),
                _fmt(item.get("reason")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="(none)")
        next_row += 1
    next_row += 1

    # Small Changes
    _write_section_label(ws, next_row, "Small Changes", width=4)
    next_row += 1
    sc_headers = ["Change", "Why", "Risk", "Priority"]
    for col_idx, h in enumerate(sc_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    next_row += 1
    changes = evaluation.get("small_changes") or []
    if changes:
        for item in changes:
            values = [
                _fmt(item.get("change")),
                _fmt(item.get("why")),
                _fmt(item.get("risk")),
                _fmt(item.get("priority")),
            ]
            for col_idx, v in enumerate(values, 1):
                c = ws.cell(row=next_row, column=col_idx, value=v)
                c.alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="(none)")

    _autosize_columns(ws, max_cols=4, max_width=70)


# ---------------------------------------------------------------------------
# Full export sheet builders
# ---------------------------------------------------------------------------


def _calculate_dots_value(total_kg: Any, bodyweight_kg: Any, sex: str) -> Any:
    if _is_blank(total_kg) or _is_blank(bodyweight_kg):
        return ""
    try:
        from analytics import calculate_dots

        return calculate_dots(float(total_kg), float(bodyweight_kg), sex)
    except Exception:
        return ""


def _epley_1rm(weight_kg: Any, reps: Any) -> float:
    try:
        weight = float(weight_kg)
        reps_val = float(reps)
    except (TypeError, ValueError):
        return 0.0
    if weight <= 0 or reps_val <= 0:
        return 0.0
    return round(weight * (1 + reps_val / 30.0), 1)


def _resolve_total_kg(payload: dict[str, Any]) -> Any:
    total = payload.get("total_kg")
    if not _is_blank(total):
        return total
    lifts = [payload.get("squat_kg"), payload.get("bench_kg"), payload.get("deadlift_kg")]
    if all(not _is_blank(lift) for lift in lifts):
        try:
            return round(sum(float(lift) for lift in lifts), 1)
        except (TypeError, ValueError):
            return ""
    return ""


def _prepare_sessions_for_export(
    sessions: list[dict[str, Any]],
    weight_log: list[dict[str, Any]],
    meta: dict[str, Any],
) -> list[dict[str, Any]]:
    ordered = sorted(_deepcopy_rows(sessions), key=lambda s: str(s.get("date", "")))

    for idx, session in enumerate(ordered):
        status = _first_non_blank(session.get("status"), "completed" if session.get("completed") else "planned")
        completed = bool(session.get("completed")) or status in ("logged", "completed")
        bodyweight, bodyweight_source = _resolve_bodyweight_for_session(ordered, idx, weight_log, meta)
        session["status"] = status
        session["completed"] = completed
        session["body_weight_kg"] = bodyweight
        session["body_weight_source"] = bodyweight_source
        session["block"] = _first_non_blank(session.get("block"), "current")
        session["phase_name"] = _extract_phase_name(session.get("phase", ""))
        if session.get("wellness") and isinstance(session["wellness"], dict):
            wellness = session["wellness"]
            session["wellness"] = {
                "sleep": wellness.get("sleep"),
                "soreness": wellness.get("soreness"),
                "mood": wellness.get("mood"),
                "stress": wellness.get("stress"),
                "energy": wellness.get("energy"),
                "recorded_at": wellness.get("recorded_at"),
            }
    return ordered


def _write_meta_sheet(
    wb: Workbook,
    meta: dict[str, Any],
    *,
    creator: str,
    program_pk: str,
    version_token: str,
    sex: str,
) -> None:
    ws = wb.active
    ws.title = "Meta"
    rows = [
        ["Program Name", meta.get("program_name", "")],
        ["Creator", creator],
        ["Program PK", program_pk],
        ["Program Version Token", version_token],
        ["Version Label", meta.get("version_label", "")],
        ["Sex", sex],
        ["Program Start", meta.get("program_start", "")],
        ["Competition Date", meta.get("comp_date", "")],
        ["Federation", meta.get("federation", "")],
        ["Practicing For", meta.get("practicing_for", "")],
        ["Weight Class (kg)", meta.get("weight_class_kg", "")],
        ["Weight Class Confirm By", meta.get("weight_class_confirm_by", "")],
        ["Current Body Weight (kg)", _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))],
        ["Current Body Weight (lb)", meta.get("current_body_weight_lb", "")],
        ["Target Squat (kg)", meta.get("target_squat_kg", "")],
        ["Target Bench (kg)", meta.get("target_bench_kg", "")],
        ["Target Deadlift (kg)", meta.get("target_dl_kg", "")],
        ["Target Total (kg)", meta.get("target_total_kg", "")],
        ["Attempt Pct - Opener", meta.get("attempt_pct", {}).get("opener", "") if isinstance(meta.get("attempt_pct"), dict) else ""],
        ["Attempt Pct - Second", meta.get("attempt_pct", {}).get("second", "") if isinstance(meta.get("attempt_pct"), dict) else ""],
        ["Attempt Pct - Third", meta.get("attempt_pct", {}).get("third", "") if isinstance(meta.get("attempt_pct"), dict) else ""],
        ["Height (cm)", meta.get("height_cm", "")],
        ["Arm Wingspan (cm)", meta.get("arm_wingspan_cm", "")],
        ["Leg Length (cm)", meta.get("leg_length_cm", "")],
        ["Equipment", meta.get("equipment", "")],
        ["Archived", meta.get("archived", "")],
        ["Archived At", meta.get("archived_at", "")],
        ["Last Updated", meta.get("updated_at", "")],
        ["Training Notes", _fmt(meta.get("training_notes", []))],
        ["Change Log", _serialize_json(meta.get("change_log", []))],
    ]
    _write_kv_sheet(ws, rows)


def _write_goals_sheet(
    wb: Workbook,
    meta: dict[str, Any],
    goals: list[dict[str, Any]],
    competitions: list[dict[str, Any]],
    sex: str,
    federation_library: dict[str, Any] | None = None,
) -> None:
    ws = wb.create_sheet("Goals")
    legacy_target_total = _resolve_total_kg({
        "squat_kg": meta.get("target_squat_kg"),
        "bench_kg": meta.get("target_bench_kg"),
        "deadlift_kg": meta.get("target_dl_kg"),
        "total_kg": meta.get("target_total_kg"),
    })
    target_bodyweight = _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))
    primary_goals = [goal.get("title", "") for goal in goals if goal.get("priority") == "primary"]
    top_rows = [
        ["Program Name", meta.get("program_name", "")],
        ["Explicit Goals Recorded", len(goals)],
        ["Primary Goals", ", ".join(primary_goals) if primary_goals else ""],
        ["Current Weight Class (kg)", meta.get("weight_class_kg", "")],
        ["Current Body Weight (kg)", target_bodyweight],
        ["Legacy Target Total (kg)", legacy_target_total],
        ["Legacy Target DOTS", _calculate_dots_value(legacy_target_total, target_bodyweight, sex)],
        ["Attempt Pct", _serialize_json(meta.get("attempt_pct", {}))],
    ]
    for row_idx, (field, value) in enumerate(top_rows, 1):
        ws.cell(row=row_idx, column=1, value=field).font = _HEADER_FONT
        ws.cell(row=row_idx, column=1).fill = _HEADER_FILL
        ws.cell(row=row_idx, column=1).alignment = _WRAP
        ws.cell(row=row_idx, column=2, value=_truncate(_fmt(value))).alignment = _WRAP

    next_row = len(top_rows) + 2
    _write_section_label(ws, next_row, "Explicit Goals", width=15)
    next_row += 1
    headers = [
        "Title", "Type", "Priority", "Strategy", "Target Date", "Competition",
        "Federation", "Standard", "Target Total", "Target DOTS", "Target IPF GL",
        "Target Weight Class", "Acceptable Classes", "Risk", "Notes",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    next_row += 1

    federations = {
        str(item.get("id")): item
        for item in (federation_library or {}).get("federations", []) or []
        if isinstance(item, dict) and item.get("id")
    }
    standards = {
        str(item.get("id")): item
        for item in (federation_library or {}).get("qualification_standards", []) or []
        if isinstance(item, dict) and item.get("id")
    }
    competitions_by_date = {
        str(comp.get("date")): comp
        for comp in competitions
        if comp.get("date")
    }

    if goals:
        priority_order = {"primary": 0, "secondary": 1, "optional": 2}
        for goal in sorted(
            goals,
            key=lambda item: (
                priority_order.get(str(item.get("priority")), 9),
                str(
                    (list(item.get("target_competition_dates") or []) or [item.get("target_competition_date") or item.get("target_date") or ""])[0]
                ),
            ),
        ):
            target_standard_ids = [str(value or "").strip() for value in list(goal.get("target_standard_ids") or []) if str(value or "").strip()]
            legacy_standard_id = str(goal.get("target_standard_id") or "").strip()
            if legacy_standard_id and legacy_standard_id not in target_standard_ids:
                target_standard_ids.append(legacy_standard_id)
            linked_standards = [standards.get(standard_id) for standard_id in target_standard_ids if standards.get(standard_id)]
            standard = linked_standards[0] if linked_standards else None
            federation = federations.get(str(goal.get("target_federation_id") or "")) or (
                federations.get(str(standard.get("federation_id"))) if isinstance(standard, dict) else None
            )
            target_competition_dates = [str(value or "").strip() for value in list(goal.get("target_competition_dates") or []) if str(value or "").strip()]
            legacy_competition_date = str(goal.get("target_competition_date") or "").strip()
            if legacy_competition_date and legacy_competition_date not in target_competition_dates:
                target_competition_dates.append(legacy_competition_date)
            competition_names = [competitions_by_date.get(target_date, {}).get("name", "") for target_date in target_competition_dates if competitions_by_date.get(target_date)]
            target_total = _first_non_blank(goal.get("target_total_kg"), (standard or {}).get("required_total_kg"))
            standard_label = "; ".join(
                f"{item.get('season_year', '')} {item.get('weight_class_kg', '')}kg {item.get('required_total_kg', '')}kg"
                for item in linked_standards
            )
            row = [
                goal.get("title", ""),
                goal.get("goal_type", ""),
                goal.get("priority", ""),
                goal.get("strategy_mode", ""),
                _first_non_blank(goal.get("target_date"), ", ".join(target_competition_dates), ""),
                ", ".join(name for name in competition_names if name),
                _first_non_blank((federation or {}).get("abbreviation"), (federation or {}).get("name"), ""),
                standard_label,
                target_total,
                goal.get("target_dots", ""),
                goal.get("target_ipf_gl", ""),
                _first_non_blank(goal.get("target_weight_class_kg"), (standard or {}).get("weight_class_kg"), ""),
                _serialize_json(goal.get("acceptable_weight_classes_kg", [])),
                goal.get("risk_tolerance", ""),
                goal.get("notes", ""),
            ]
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=next_row, column=col_idx, value=_truncate(_fmt(value))).alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="No explicit goals recorded. Legacy targets are shown in the header rows.")
        next_row += 1

    next_row += 2
    _write_section_label(ws, next_row, "Competition Targets", width=13)
    next_row += 1
    federation_names = {
        str(item.get("id")): _first_non_blank(item.get("abbreviation"), item.get("name"), "")
        for item in (federation_library or {}).get("federations", []) or []
        if isinstance(item, dict)
    }

    comp_headers = [
        "Name", "Date", "Status", "Host Federation", "Counts Toward", "Weight Class", "Body Weight",
        "Target Squat", "Target Bench", "Target Deadlift", "Target Total", "Target DOTS", "Notes",
    ]
    for col_idx, header in enumerate(comp_headers, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    next_row += 1

    if competitions:
        for comp in sorted(competitions, key=lambda c: str(c.get("date", ""))):
            targets = comp.get("targets") or {}
            resolved_bw = _first_non_blank(comp.get("body_weight_kg"), _resolve_bodyweight_for_date(comp.get("date"), [], [], meta)[0])
            target_total = _resolve_total_kg(targets)
            row = [
                comp.get("name", ""),
                comp.get("date", ""),
                comp.get("status", ""),
                _first_non_blank(
                    federation_names.get(str(comp.get("federation_id") or "")),
                    comp.get("federation", ""),
                ),
                ", ".join(
                    federation_names.get(str(federation_id), str(federation_id))
                    for federation_id in (comp.get("counts_toward_federation_ids") or [])
                ),
                comp.get("weight_class_kg", ""),
                resolved_bw,
                targets.get("squat_kg", ""),
                targets.get("bench_kg", ""),
                targets.get("deadlift_kg", ""),
                target_total,
                _calculate_dots_value(target_total, resolved_bw, sex),
                comp.get("notes", ""),
            ]
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=next_row, column=col_idx, value=_truncate(_fmt(value))).alignment = _WRAP
            next_row += 1
    else:
        ws.cell(row=next_row, column=1, value="No competitions recorded.")

    _autosize_columns(ws, max_cols=max(len(headers), len(comp_headers)))


def _write_federation_standards_sheet(wb: Workbook, federation_library: dict[str, Any] | None) -> None:
    ws = wb.create_sheet("Federation Standards")
    library = federation_library or {}
    federations = list(library.get("federations") or [])
    standards = list(library.get("qualification_standards") or [])
    federation_names = {
        str(item.get("id")): _first_non_blank(item.get("abbreviation"), item.get("name"), "")
        for item in federations
        if isinstance(item, dict)
    }

    _write_section_label(ws, 1, "Federations", width=5)
    fed_headers = ["Name", "Abbreviation", "Region", "Status", "Notes"]
    for col_idx, header in enumerate(fed_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP

    row_idx = 3
    if federations:
        for item in federations:
            values = [
                item.get("name", ""),
                item.get("abbreviation", ""),
                item.get("region", ""),
                item.get("status", ""),
                item.get("notes", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=_truncate(_fmt(value))).alignment = _WRAP
            row_idx += 1
    else:
        ws.cell(row=row_idx, column=1, value="No federations recorded.")
        row_idx += 1

    row_idx += 2
    _write_section_label(ws, row_idx, "Qualification Standards", width=13)
    row_idx += 1
    std_headers = [
        "Federation", "Competition", "Season", "Sex", "Equipment", "Event", "Age Class",
        "Division", "Weight Class", "Required Total", "Qualifying Start", "Qualifying End",
        "Source Label", "Source URL", "Status",
    ]
    for col_idx, header in enumerate(std_headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    row_idx += 1

    if standards:
        for item in sorted(standards, key=lambda entry: (str(entry.get("status", "")), -int(_num(entry.get("season_year")) or 0), _num(entry.get("weight_class_kg")))):
            values = [
                federation_names.get(str(item.get("federation_id")), ""),
                item.get("competition_name", ""),
                item.get("season_year", ""),
                item.get("sex", ""),
                item.get("equipment", ""),
                item.get("event", ""),
                item.get("age_class", ""),
                item.get("division", ""),
                item.get("weight_class_kg", ""),
                item.get("required_total_kg", ""),
                item.get("qualifying_start_date", ""),
                item.get("qualifying_end_date", ""),
                item.get("source_label", ""),
                item.get("source_url", ""),
                item.get("status", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=_truncate(_fmt(value))).alignment = _WRAP
            row_idx += 1
    else:
        ws.cell(row=row_idx, column=1, value="No qualification standards recorded.")

    _autosize_columns(ws, max_cols=len(std_headers))


def _write_current_maxes_sheet(wb: Workbook, current_maxes: dict[str, Any]) -> None:
    ws = wb.create_sheet("Current Maxes")
    rows: list[list[Any]] = []
    for lift, val in current_maxes.items():
        if lift == "_note":
            continue
        rows.append([lift.replace("_", " ").title(), val])
    if not rows:
        rows.append(["—", "No current maxes recorded"])
    _write_sheet(ws, ["Lift", "Current Max (kg)"], rows)


def _write_max_history_sheet(
    wb: Workbook,
    max_history: list[dict[str, Any]],
    sex: str,
    meta: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Max History")
    headers = ["Date", "Squat", "Bench", "Deadlift", "Total", "Bodyweight", "DOTS", "Context"]
    rows = []
    for entry in sorted(max_history, key=lambda e: str(e.get("date", ""))):
        bw = _first_non_blank(entry.get("bodyweight_kg"), meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"))
        total = _first_non_blank(entry.get("total_kg"), 0)
        rows.append([
            entry.get("date", ""),
            entry.get("squat_kg", ""),
            entry.get("bench_kg", ""),
            entry.get("deadlift_kg", ""),
            total,
            bw,
            _calculate_dots_value(total, bw, sex),
            entry.get("context", ""),
        ])
    if not rows:
        rows.append(["—", "", "", "", "", "", "", "No max history recorded"])
    _write_sheet(ws, headers, rows)


def _write_phases_sheet(wb: Workbook, phases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Phases")
    headers = ["Phase", "Block", "Start Week", "End Week", "Intent", "Target RPE Min", "Target RPE Max", "Days/Week", "Notes"]
    rows = []
    for phase in phases:
        rows.append([
            phase.get("name", ""),
            _first_non_blank(phase.get("block"), "current"),
            phase.get("start_week", ""),
            phase.get("end_week", ""),
            phase.get("intent", ""),
            phase.get("target_rpe_min", ""),
            phase.get("target_rpe_max", ""),
            phase.get("days_per_week", ""),
            phase.get("notes", ""),
        ])
    if not rows:
        rows.append(["—", "current", "", "", "No phases recorded", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={5: 28, 9: 40})


def _write_sessions_sheet(wb: Workbook, sessions: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Sessions")
    headers = [
        "ID", "Date", "Week", "Week #", "Block", "Phase", "Day", "Status", "Completed",
        "Body Weight (kg)", "Body Weight Source", "Session RPE", "Sleep", "Soreness", "Mood",
        "Stress", "Energy", "Wellness Recorded At", "Notes",
    ]
    rows = []
    for session in sessions:
        wellness = session.get("wellness") or {}
        rows.append([
            session.get("id", ""),
            session.get("date", ""),
            session.get("week", ""),
            session.get("week_number", ""),
            session.get("block", ""),
            _extract_phase_name(session.get("phase", "")),
            session.get("day", ""),
            session.get("status", ""),
            session.get("completed", ""),
            session.get("body_weight_kg", ""),
            session.get("body_weight_source", ""),
            session.get("session_rpe", ""),
            wellness.get("sleep", ""),
            wellness.get("soreness", ""),
            wellness.get("mood", ""),
            wellness.get("stress", ""),
            wellness.get("energy", ""),
            wellness.get("recorded_at", ""),
            session.get("session_notes", ""),
        ])
    if not rows:
        rows.append(["—", "", "", "", "", "", "", "No sessions recorded", "", "", "", "", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={19: 42})


def _write_planned_exercises_sheet(wb: Workbook, sessions: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Planned Exercises")
    headers = ["Date", "Week", "Block", "Phase", "Exercise", "Sets", "Reps", "Weight (kg)", "Load Source", "RPE Target", "Notes", "Session Status"]
    rows = []
    for session in sessions:
        planned = session.get("planned_exercises") or []
        for ex in planned:
            rows.append([
                session.get("date", ""),
                session.get("week_number", ""),
                session.get("block", ""),
                session.get("phase_name", ""),
                ex.get("name", ""),
                ex.get("sets", ""),
                ex.get("reps", ""),
                ex.get("kg", ""),
                ex.get("load_source", ""),
                ex.get("rpe_target", ""),
                ex.get("notes", ""),
                session.get("status", ""),
            ])
    if not rows:
        rows.append(["—", "", "", "", "No planned exercises recorded", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows)


def _write_exercises_sheet(wb: Workbook, sessions: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Exercises")
    headers = [
        "Date", "Week", "Block", "Phase", "Exercise", "Sets", "Reps", "Weight (kg)",
        "Failed", "Failed Sets", "RPE", "Notes", "Volume",
    ]
    rows = []
    for session in sessions:
        for ex in session.get("exercises", []) or []:
            sets_f = float(ex.get("sets") or 0)
            reps_f = float(ex.get("reps") or 0)
            kg_f = float(ex.get("kg") or 0)
            failed_sets = ex.get("failed_sets")
            if isinstance(failed_sets, list):
                failed_sets_val = "; ".join("1" if bool(v) else "0" for v in failed_sets)
            else:
                failed_sets_val = ""
            rows.append([
                session.get("date", ""),
                session.get("week_number", ""),
                session.get("block", ""),
                session.get("phase_name", ""),
                ex.get("name", ""),
                ex.get("sets", ""),
                ex.get("reps", ""),
                ex.get("kg", ""),
                ex.get("failed", False),
                failed_sets_val,
                ex.get("rpe", ""),
                ex.get("notes", ""),
                round(sets_f * reps_f * kg_f, 1),
            ])
    if not rows:
        rows.append(["—", "", "", "", "No exercises recorded", "", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={12: 40})


def _write_competitions_sheet(
    wb: Workbook,
    competitions: list[dict[str, Any]],
    sessions: list[dict[str, Any]],
    weight_log: list[dict[str, Any]],
    sex: str,
    meta: dict[str, Any],
    federation_library: dict[str, Any] | None = None,
) -> None:
    ws = wb.create_sheet("Competitions")
    federation_names = {
        str(item.get("id")): _first_non_blank(item.get("abbreviation"), item.get("name"), "")
        for item in (federation_library or {}).get("federations", []) or []
        if isinstance(item, dict)
    }
    headers = [
        "Name", "Date", "Status", "Federation", "Counts Toward", "Location", "Hotel Required", "Weight Class",
        "Body Weight", "Body Weight Source", "Target Squat", "Target Bench", "Target Deadlift",
        "Target Total", "Target DOTS", "Result Squat", "Result Bench", "Result Deadlift",
        "Result Total", "Result DOTS", "Place", "Decision Date", "Projection Snapshot Date",
        "Notes", "Between Comp Plan", "Comp Day Protocol",
    ]
    rows = []
    for comp in sorted(competitions, key=lambda c: str(c.get("date", ""))):
        targets = comp.get("targets") or {}
        results = comp.get("results") or {}
        bodyweight = _first_non_blank(comp.get("body_weight_kg"), _resolve_bodyweight_for_date(comp.get("date"), sessions, weight_log, meta)[0])
        bw_source = "competition" if not _is_blank(comp.get("body_weight_kg")) else _resolve_bodyweight_for_date(comp.get("date"), sessions, weight_log, meta)[1]
        target_total = _resolve_total_kg(targets)
        result_total = _resolve_total_kg(results)
        rows.append([
            comp.get("name", ""),
            comp.get("date", ""),
            comp.get("status", ""),
            _first_non_blank(
                federation_names.get(str(comp.get("federation_id") or "")),
                comp.get("federation", ""),
            ),
            ", ".join(
                federation_names.get(str(federation_id), str(federation_id))
                for federation_id in (comp.get("counts_toward_federation_ids") or [])
            ),
            comp.get("location", ""),
            comp.get("hotel_required", ""),
            comp.get("weight_class_kg", ""),
            bodyweight,
            bw_source,
            targets.get("squat_kg", ""),
            targets.get("bench_kg", ""),
            targets.get("deadlift_kg", ""),
            target_total,
            _calculate_dots_value(target_total, bodyweight, sex),
            results.get("squat_kg", ""),
            results.get("bench_kg", ""),
            results.get("deadlift_kg", ""),
            result_total,
            _calculate_dots_value(result_total, bodyweight, sex) if comp.get("status") == "completed" else "",
            comp.get("place", ""),
            comp.get("decision_date", ""),
            comp.get("projection_snapshot_date", ""),
            comp.get("notes", ""),
            _serialize_json(comp.get("between_comp_plan", {})),
            _serialize_json(comp.get("comp_day_protocol", {})),
        ])
    if not rows:
        rows.append(["—", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "No competitions recorded", "", ""])
    _write_sheet(ws, headers, rows, col_widths={23: 40, 24: 36, 25: 36})


def _write_biometrics_sheet(wb: Workbook, diet_notes: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Biometrics")
    headers = ["Date", "Notes", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)", "Sleep (hrs)", "Water", "Water Unit", "Consistent"]
    rows = []
    for note in sorted(diet_notes, key=lambda n: str(n.get("date", ""))):
        rows.append([
            note.get("date", ""),
            note.get("notes", ""),
            note.get("avg_daily_calories", ""),
            note.get("avg_protein_g", ""),
            note.get("avg_carb_g", ""),
            note.get("avg_fat_g", ""),
            note.get("avg_sleep_hours", ""),
            note.get("water_intake", ""),
            note.get("water_unit", ""),
            note.get("consistent", ""),
        ])
    if not rows:
        rows.append(["—", "No biometrics entries recorded", "", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={2: 44})


def _write_weight_log_sheet(wb: Workbook, weight_log: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Weight Log")
    headers = ["Date", "Kg", "Delta From Previous", "Notes"]
    rows = []
    sorted_log = sorted(weight_log, key=lambda e: str(e.get("date", "")))
    previous = None
    for entry in sorted_log:
        kg = entry.get("kg", "")
        delta = ""
        if previous is not None and not _is_blank(kg):
            try:
                delta = round(float(kg) - float(previous), 1)
            except (TypeError, ValueError):
                delta = ""
        rows.append([entry.get("date", ""), kg, delta, entry.get("notes", "")])
        previous = kg if not _is_blank(kg) else previous
    if not rows:
        rows.append(["—", "", "", "No weight log recorded"])
    _write_sheet(ws, headers, rows)


def _write_supplements_sheet(wb: Workbook, supplements: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Supplements")
    headers = ["Name", "Dose"]
    rows = [[supp.get("name", ""), supp.get("dose", "")] for supp in supplements]
    if not rows:
        rows.append(["—", "No supplements recorded"])
    _write_sheet(ws, headers, rows)


def _write_supplement_phases_sheet(wb: Workbook, supplement_phases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Supplement Phases")
    headers = ["Phase", "Phase Name", "Block", "Start Week", "End Week", "Supplement", "Dose", "Item Notes", "Phase Notes", "Peak Week Protocol"]
    rows = []
    for phase in sorted(supplement_phases, key=lambda p: int(p.get("phase", 0) or 0)):
        items = phase.get("items") or []
        if not items:
            rows.append([
                phase.get("phase", ""),
                phase.get("phase_name", ""),
                phase.get("block", ""),
                phase.get("start_week", ""),
                phase.get("end_week", ""),
                "",
                "",
                "",
                phase.get("notes", ""),
                _serialize_json(phase.get("peak_week_protocol", {})),
            ])
            continue
        for item in items:
            rows.append([
                phase.get("phase", ""),
                phase.get("phase_name", ""),
                phase.get("block", ""),
                phase.get("start_week", ""),
                phase.get("end_week", ""),
                item.get("name", ""),
                item.get("dose", ""),
                item.get("notes", ""),
                phase.get("notes", ""),
                _serialize_json(phase.get("peak_week_protocol", {})),
            ])
    if not rows:
        rows.append(["—", "", "", "", "", "No supplement phases recorded", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={9: 34, 10: 36})


def _write_exercise_glossary_sheet(wb: Workbook, glossary: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Exercise Glossary")
    headers = [
        "ID", "Name", "Category", "Fatigue Category", "Primary Muscles", "Secondary Muscles", "Tertiary Muscles",
        "Equipment", "Cues", "Notes", "Video URL", "Fatigue Axial", "Fatigue Neural",
        "Fatigue Peripheral", "Fatigue Systemic", "Fatigue Source", "Fatigue Reasoning",
        "E1RM Value", "E1RM Method", "E1RM Basis", "E1RM Confidence", "E1RM Updated At", "Archived",
    ]
    rows = []
    for ex in glossary:
        fatigue = ex.get("fatigue_profile") or {}
        e1rm = ex.get("e1rm_estimate") or {}
        rows.append([
            ex.get("id", ""),
            ex.get("name", ""),
            ex.get("category", ""),
            ex.get("fatigue_category", ""),
            _fmt(ex.get("primary_muscles", [])),
            _fmt(ex.get("secondary_muscles", [])),
            _fmt(ex.get("tertiary_muscles", [])),
            ex.get("equipment", ""),
            _fmt(ex.get("cues", [])),
            ex.get("notes", ""),
            ex.get("video_url", ""),
            fatigue.get("axial", ""),
            fatigue.get("neural", ""),
            fatigue.get("peripheral", ""),
            fatigue.get("systemic", ""),
            ex.get("fatigue_profile_source", ""),
            ex.get("fatigue_profile_reasoning", ""),
            e1rm.get("value_kg", ""),
            e1rm.get("method", ""),
            e1rm.get("basis", ""),
            e1rm.get("confidence", ""),
            e1rm.get("set_at", ""),
            ex.get("archived", ""),
        ])
    if not rows:
        rows.append(["—", "", "", "", "", "", "", "", "", "No glossary entries recorded", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={2: 34, 9: 34, 10: 40, 17: 38})


def _write_videos_sheet(wb: Workbook, sessions: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Videos")
    headers = ["Session Date", "Block", "Exercise", "Set #", "Notes", "Uploaded At", "Video URL", "Thumbnail URL", "Thumbnail Status"]
    rows = []
    for session in sessions:
        for video in session.get("videos") or []:
            rows.append([
                session.get("date", ""),
                session.get("block", ""),
                video.get("exercise_name", ""),
                video.get("set_number", ""),
                video.get("notes", ""),
                video.get("uploaded_at", ""),
                video.get("video_url", ""),
                video.get("thumbnail_url", ""),
                video.get("thumbnail_status", ""),
            ])
    if not rows:
        rows.append(["—", "", "No videos recorded", "", "", "", "", "", ""])
    _write_sheet(ws, headers, rows, col_widths={5: 34, 7: 42, 8: 42})


def _write_notes_sheet(
    wb: Workbook,
    meta: dict[str, Any],
    phases: list[dict[str, Any]],
    sessions: list[dict[str, Any]],
    competitions: list[dict[str, Any]],
    diet_notes: list[dict[str, Any]],
    supplement_phases: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Notes")
    headers = ["Source", "Date", "Title", "Notes"]
    rows: list[list[Any]] = []

    for note in _safe_list(meta.get("training_notes")):
        rows.append(["Meta", meta.get("updated_at", ""), "Training Note", note])

    for phase in phases:
        if phase.get("notes"):
            rows.append(["Phase", f"W{phase.get('start_week', '')}-W{phase.get('end_week', '')}", phase.get("name", ""), phase.get("notes", "")])

    for session in sessions:
        if session.get("session_notes"):
            rows.append(["Session", session.get("date", ""), session.get("day", ""), session.get("session_notes", "")])

    for comp in competitions:
        if comp.get("notes"):
            rows.append(["Competition", comp.get("date", ""), comp.get("name", ""), comp.get("notes", "")])

    for note in diet_notes:
        if note.get("notes"):
            rows.append(["Biometrics", note.get("date", ""), "Diet Note", note.get("notes", "")])

    for phase in supplement_phases:
        if phase.get("notes"):
            rows.append(["Supplement Phase", phase.get("phase_name", ""), f"Phase {phase.get('phase', '')}", phase.get("notes", "")])
        for item in phase.get("items") or []:
            if item.get("notes"):
                rows.append(["Supplement Item", phase.get("phase_name", ""), item.get("name", ""), item.get("notes", "")])

    if not rows:
        rows.append(["—", "", "No notes recorded", ""])
    _write_sheet(ws, headers, rows, col_widths={4: 56})


def _write_breaks_and_prefs_sheet(
    wb: Workbook,
    breaks: list[dict[str, Any]],
    operator_prefs: dict[str, Any],
    meta: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Breaks & Prefs")
    rows = [
        ["Archived", meta.get("archived", "")],
        ["Archived At", meta.get("archived_at", "")],
        ["Operator Prefs", _serialize_json(operator_prefs)],
    ]
    if breaks:
        rows.append(["Breaks", _serialize_json(breaks)])
    else:
        rows.append(["Breaks", "[]"])
    _write_kv_sheet(ws, rows)


def _write_trends_sheet(
    wb: Workbook,
    sessions: list[dict[str, Any]],
    weight_log: list[dict[str, Any]],
    diet_notes: list[dict[str, Any]],
    sex: str,
    meta: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Trends")
    row = 1

    # Bodyweight trend
    _write_section_label(ws, row, "Bodyweight Trend", width=6)
    row += 1
    bw_rows: list[list[Any]] = []
    bw_points: list[tuple[str, Any]] = []
    source_log = sorted(weight_log, key=lambda e: str(e.get("date", "")))
    if source_log:
        for entry in source_log:
            if not _is_blank(entry.get("kg")):
                bw_points.append((str(entry.get("date", "")), entry.get("kg")))
    else:
        for session in sessions:
            if session.get("block", "current") != "current" or not session.get("completed"):
                continue
            if not _is_blank(session.get("body_weight_kg")):
                bw_points.append((str(session.get("date", "")), session.get("body_weight_kg")))
    bw_points.sort(key=lambda item: item[0])
    if bw_points:
        latest = _num(bw_points[-1][1])
        oldest = _num(bw_points[0][1])
        change = round(latest - oldest, 1)
        direction = "gain" if change > 0.25 else "loss" if change < -0.25 else "stable"
        bw_rows = [
            ["Latest", latest],
            ["Oldest", oldest],
            ["Change", change],
            ["Direction", direction],
            ["Entries", len(bw_points)],
        ]
    else:
        bw_rows = [["Status", "No bodyweight data"]]
    for offset, (field, value) in enumerate(bw_rows, row):
        ws.cell(row=offset, column=1, value=field).font = _HEADER_FONT
        ws.cell(row=offset, column=1).fill = _HEADER_FILL
        ws.cell(row=offset, column=2, value=value)
    row += len(bw_rows) + 1

    if bw_points:
        headers = ["Date", "Bodyweight (kg)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        row += 1
        for date_str, kg in bw_points[-12:]:
            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=_num(kg))
            row += 1
        row += 1

    # Nutrition trend
    _write_section_label(ws, row, "Nutrition Trend", width=6)
    row += 1
    if diet_notes:
        metrics = {
            "Calories": [n.get("avg_daily_calories") for n in diet_notes if not _is_blank(n.get("avg_daily_calories"))],
            "Protein": [n.get("avg_protein_g") for n in diet_notes if not _is_blank(n.get("avg_protein_g"))],
            "Carbs": [n.get("avg_carb_g") for n in diet_notes if not _is_blank(n.get("avg_carb_g"))],
            "Fat": [n.get("avg_fat_g") for n in diet_notes if not _is_blank(n.get("avg_fat_g"))],
            "Sleep": [n.get("avg_sleep_hours") for n in diet_notes if not _is_blank(n.get("avg_sleep_hours"))],
            "Water": [n.get("water_intake") for n in diet_notes if not _is_blank(n.get("water_intake"))],
        }
        nutrition_rows = [["Consistency %", round((sum(1 for n in diet_notes if n.get("consistent")) / len(diet_notes)) * 100, 1)]]
        for label, values in metrics.items():
            if values:
                avg_value = round(sum(float(v) for v in values) / len(values), 1)
            else:
                avg_value = ""
            nutrition_rows.append([label, avg_value])
        for offset, (field, value) in enumerate(nutrition_rows, row):
            ws.cell(row=offset, column=1, value=field).font = _HEADER_FONT
            ws.cell(row=offset, column=1).fill = _HEADER_FILL
            ws.cell(row=offset, column=2, value=value)
        row += len(nutrition_rows) + 1

        headers = ["Date", "Notes", "Calories", "Protein", "Carbs", "Fat", "Sleep", "Water", "Water Unit", "Consistent"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        row += 1
        for note in sorted(diet_notes, key=lambda n: str(n.get("date", ""))):
            ws.cell(row=row, column=1, value=note.get("date", ""))
            ws.cell(row=row, column=2, value=note.get("notes", ""))
            ws.cell(row=row, column=3, value=note.get("avg_daily_calories", ""))
            ws.cell(row=row, column=4, value=note.get("avg_protein_g", ""))
            ws.cell(row=row, column=5, value=note.get("avg_carb_g", ""))
            ws.cell(row=row, column=6, value=note.get("avg_fat_g", ""))
            ws.cell(row=row, column=7, value=note.get("avg_sleep_hours", ""))
            ws.cell(row=row, column=8, value=note.get("water_intake", ""))
            ws.cell(row=row, column=9, value=note.get("water_unit", ""))
            ws.cell(row=row, column=10, value=note.get("consistent", ""))
            row += 1
        row += 1
    else:
        ws.cell(row=row, column=1, value="No nutrition data recorded")
        row += 2

    # DOTS trend
    _write_section_label(ws, row, "DOTS Trend", width=6)
    row += 1
    trend_rows = _build_dots_trend_rows(sessions, weight_log, sex, meta)
    headers = ["Week", "Squat", "Bench", "Deadlift", "Total", "Bodyweight", "DOTS"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row += 1
    if trend_rows:
        for item in trend_rows:
            ws.cell(row=row, column=1, value=item.get("week", ""))
            ws.cell(row=row, column=2, value=item.get("squat", ""))
            ws.cell(row=row, column=3, value=item.get("bench", ""))
            ws.cell(row=row, column=4, value=item.get("deadlift", ""))
            ws.cell(row=row, column=5, value=item.get("total", ""))
            ws.cell(row=row, column=6, value=item.get("bodyweight", ""))
            ws.cell(row=row, column=7, value=item.get("dots", ""))
            row += 1
    else:
        ws.cell(row=row, column=1, value="No DOTS trend data available")

    _autosize_columns(ws, max_cols=10, max_width=60)


def _build_dots_trend_rows(
    sessions: list[dict[str, Any]],
    weight_log: list[dict[str, Any]],
    sex: str,
    meta: dict[str, Any],
) -> list[dict[str, Any]]:
    current_block_sessions = [
        s for s in sessions
        if s.get("block", "current") == "current" and (s.get("completed") or s.get("status") in ("logged", "completed"))
    ]
    by_week: dict[Any, dict[str, Any]] = {}
    for session in current_block_sessions:
        week = session.get("week_number") or session.get("week") or 0
        bucket = by_week.setdefault(week, {"squat": 0.0, "bench": 0.0, "deadlift": 0.0, "bodyweight": 0.0})
        bodyweight = session.get("body_weight_kg")
        if not _is_blank(bodyweight) and float(bodyweight) > float(bucket["bodyweight"]):
            bucket["bodyweight"] = float(bodyweight)
        for ex in session.get("exercises") or []:
            name = str(ex.get("name", "")).lower()
            weight = ex.get("kg")
            reps = ex.get("reps")
            if _is_blank(weight) or _is_blank(reps):
                continue
            e1rm = _epley_1rm(weight, reps)
            if "squat" in name and "hack" not in name and "split" not in name:
                bucket["squat"] = max(bucket["squat"], e1rm)
            elif "bench" in name:
                bucket["bench"] = max(bucket["bench"], e1rm)
            elif "deadlift" in name and "romanian" not in name and "rdl" not in name:
                bucket["deadlift"] = max(bucket["deadlift"], e1rm)

    rows: list[dict[str, Any]] = []
    for week in sorted(by_week.keys(), key=lambda w: int(w) if str(w).isdigit() else str(w)):
        bucket = by_week[week]
        total = round(bucket["squat"] + bucket["bench"] + bucket["deadlift"], 1)
        bodyweight = bucket["bodyweight"] or _first_non_blank(meta.get("current_body_weight_kg"), meta.get("bodyweight_kg"), _nearest_weight_log_value(None, weight_log))
        dots = _calculate_dots_value(total, bodyweight, sex) if total and bodyweight else ""
        rows.append({
            "week": week,
            "squat": round(bucket["squat"], 1) if bucket["squat"] else "",
            "bench": round(bucket["bench"], 1) if bucket["bench"] else "",
            "deadlift": round(bucket["deadlift"], 1) if bucket["deadlift"] else "",
            "total": total if total else "",
            "bodyweight": _num(bodyweight) if bodyweight else "",
            "dots": dots,
        })
    return rows
