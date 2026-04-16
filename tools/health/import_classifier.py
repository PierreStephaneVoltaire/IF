"""Deterministic pre-classification and extraction for import files.

Identifies if a file is a Template (date-free) or a Session Import (dated)
before falling back to AI classification.
"""
from __future__ import annotations

import csv
import hashlib
import io
import logging
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)

def file_hash(file_bytes: bytes) -> str:
    """Return a short sha256 hash prefix for the file."""
    h = hashlib.sha256(file_bytes).hexdigest()
    return f"sha256:{h[:16]}"

def extract_xlsx(file_bytes: bytes) -> tuple[list[dict[str, Any]], str]:
    """Extract rows from the first non-empty sheet of an XLSX file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    
    # Find first sheet with data
    sheet = wb.active
    for sname in wb.sheetnames:
        if wb[sname].max_row > 1:
            sheet = wb[sname]
            break
            
    rows = []
    headers = [str(cell.value).strip() if cell.value else f"col_{i}" for i, cell in enumerate(sheet[1])]
    
    for row_cells in sheet.iter_rows(min_row=2):
        row_dict = {}
        has_data = False
        for i, cell in enumerate(row_cells):
            if i < len(headers):
                val = cell.value
                row_dict[headers[i]] = val
                if val is not None:
                    has_data = True
        if has_data:
            rows.append(row_dict)
            
    return rows, sheet.title

def extract_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    """Extract rows from a CSV file."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader if any(row.values())]

def preclassify_rows(rows: list[dict[str, Any]]) -> Optional[str]:
    """Identify 'template' vs 'session_import' via heuristics.
    
    Returns 'template', 'session_import', or None if ambiguous.
    """
    if not rows:
        return None
        
    # Heuristic 1: Look for date-like columns
    date_cols = [k for k in rows[0].keys() if "date" in k.lower()]
    has_actual_dates = False
    for row in rows[:20]: # Check first 20 rows
        for col in date_cols:
            val = row.get(col)
            if val and (isinstance(val, (datetime, date)) or (isinstance(val, str) and "-" in val and len(val) >= 8)):
                has_actual_dates = True
                break
        if has_actual_dates: break

    if has_actual_dates:
        return "session_import"

    # Heuristic 2: Week/Day without dates
    week_cols = [k for k in rows[0].keys() if "week" in k.lower()]
    if week_cols:
        # If we have "Week 1", "Week 2" etc but no dates, it's likely a template
        return "template"

    # Heuristic 3: RPE/% presence
    load_cols = [k for k in rows[0].keys() if any(x in k.lower() for x in ["rpe", "percentage", "%", "target"])]
    if load_cols and not has_actual_dates:
        return "template"

    return None

from datetime import datetime, date
