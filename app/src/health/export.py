"""Excel export builder for training programs.

Uses openpyxl to produce a multi-sheet .xlsx workbook from a program dict.
"""
from __future__ import annotations

import os
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Shared styles
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")

# openpyxl cell content limit
_CELL_CHAR_LIMIT = 32000


def _num(v: Any) -> Any:
    """Coerce a value to an Excel-compatible scalar."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (dict, list)):
        return str(v)
    return v


def _extract_phase_name(phase: Any) -> str:
    """Extract phase name from either a string or a phase dict."""
    if isinstance(phase, dict):
        return phase.get("name", "")
    return phase if isinstance(phase, str) else str(phase)


def _truncate(val: Any, limit: int = _CELL_CHAR_LIMIT) -> Any:
    if isinstance(val, str) and len(val) > limit:
        return val[: limit - 30] + "... [truncated]"
    return val


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], col_widths: dict[int, int] | None = None):
    """Write a header row + data rows to a worksheet."""
    # Header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP

    # Data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=_truncate(_num(value)))

    # Column widths
    max_col = len(headers)
    for col_idx in range(1, max_col + 1):
        if col_widths and col_idx in col_widths:
            width = col_widths[col_idx]
        else:
            # Auto-width from header + data
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            width = min(max_len + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"


def build_program_xlsx(program: dict, out_path: str) -> str:
    """Build a multi-sheet Excel workbook from a program dict.

    Args:
        program: Full program dict from ProgramStore.
        out_path: Absolute path to write the .xlsx file.

    Returns:
        Absolute path to the written file.
    """
    wb = Workbook()

    meta = program.get("meta", {})
    phases = program.get("phases", [])
    sessions = program.get("sessions", [])
    competitions = program.get("competitions", [])
    maxes = program.get("current_maxes", {})

    # ---- Sheet 1: Meta ----
    ws_meta = wb.active
    ws_meta.title = "Meta"
    meta_rows = [
        ["Program Name", meta.get("program_name", "")],
        ["Creator", meta.get("creator", "")],
        ["Program Start", meta.get("program_start", "")],
        ["Competition Date", meta.get("comp_date", "")],
        ["Target Total (kg)", meta.get("target_total_kg", "")],
        ["Weight Class (kg)", meta.get("weight_class_kg", "")],
        ["Equipment", meta.get("equipment", "")],
        ["Last Updated", meta.get("updated_at", "")],
    ]
    _write_sheet(ws_meta, ["Field", "Value"], meta_rows)

    # ---- Sheet 2: Current Maxes ----
    ws_maxes = wb.create_sheet("Current Maxes")
    maxes_rows = []
    for lift, val in maxes.items():
        maxes_rows.append([lift.replace("_", " ").title(), val])
    _write_sheet(ws_maxes, ["Lift", "Current Max (kg)"], maxes_rows)

    # ---- Sheet 3: Phases ----
    ws_phases = wb.create_sheet("Phases")
    phase_rows = []
    for phase in phases:
        phase_rows.append([
            phase.get("name", ""),
            phase.get("start_week", ""),
            phase.get("end_week", ""),
            phase.get("intent", ""),
            phase.get("target_rpe", ""),
            phase.get("days_per_week", ""),
        ])
    _write_sheet(
        ws_phases,
        ["Phase", "Start Week", "End Week", "Intent", "Target RPE", "Days/Week"],
        phase_rows,
    )

    # ---- Sheet 4: Sessions ----
    ws_sessions = wb.create_sheet("Sessions")
    session_rows = []
    for s in sessions:
        session_rows.append([
            s.get("date", ""),
            s.get("week", ""),
            s.get("week_number", ""),
            _extract_phase_name(s.get("phase", "")),
            s.get("day", ""),
            s.get("completed", ""),
            s.get("body_weight_kg", ""),
            s.get("session_rpe", ""),
            s.get("session_notes", ""),
        ])
    _write_sheet(
        ws_sessions,
        ["Date", "Week", "Week #", "Phase", "Day", "Completed", "Body Weight (kg)", "Session RPE", "Notes"],
        session_rows,
    )

    # ---- Sheet 5: Exercises (flattened) ----
    ws_exercises = wb.create_sheet("Exercises")
    ex_rows = []
    for s in sessions:
        s_date = s.get("date", "")
        s_week = s.get("week", "")
        s_phase = _extract_phase_name(s.get("phase", ""))
        for ex in s.get("exercises", []):
            sets = ex.get("sets", 0) or 0
            reps = ex.get("reps", 0) or 0
            kg = ex.get("kg")
            try:
                kg_f = float(kg) if kg is not None else 0
            except (ValueError, TypeError):
                kg_f = 0
            volume = sets * reps * kg_f
            ex_rows.append([
                s_date,
                s_week,
                s_phase,
                ex.get("name", ""),
                sets,
                reps,
                ex.get("kg", ""),
                ex.get("rpe", ""),
                round(volume, 1),
            ])
    _write_sheet(
        ws_exercises,
        ["Date", "Week", "Phase", "Exercise", "Sets", "Reps", "Weight (kg)", "RPE", "Volume"],
        ex_rows,
    )

    # ---- Sheet 6: Competitions ----
    ws_comps = wb.create_sheet("Competitions")
    comp_rows = []
    for c in competitions:
        attempts = c.get("attempts", {})
        comp_rows.append([
            c.get("name", ""),
            c.get("date", ""),
            c.get("federation", ""),
            c.get("weight_class_kg", ""),
            c.get("equipment", ""),
            attempts.get("squat", {}).get("best", ""),
            attempts.get("bench", {}).get("best", ""),
            attempts.get("deadlift", {}).get("best", ""),
            attempts.get("total", ""),
            c.get("dots", ""),
            c.get("place", ""),
        ])
    _write_sheet(
        ws_comps,
        ["Meet", "Date", "Federation", "Weight Class", "Equipment",
         "Squat", "Bench", "Deadlift", "Total", "DOTS", "Place"],
        comp_rows,
    )

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
